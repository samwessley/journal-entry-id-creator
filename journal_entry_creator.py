#!/usr/bin/env python3
"""
Journal Entry ID Creator Script

This script reads journal lines from an Excel file and groups them into balanced
journal entries where:
1. Each journal entry has debits = credits (balanced)
2. All journal lines in an entry have the same posted date
3. Journal lines are grouped by combinations of posted date + optional fields

Algorithm:
1. Read all journal lines from Excel
2. Group lines by Posted Date
3. Within each date group, try different combinations of optional fields
4. For each combination, check if the group balances (debits = credits)
5. Assign unique Journal Entry IDs to balanced groups
6. Output results with Journal ID column added
"""

import pandas as pd
import numpy as np
from decimal import Decimal, ROUND_HALF_UP
from openpyxl import load_workbook
from itertools import combinations, product
from collections import defaultdict
import uuid
import argparse
from pathlib import Path

class JournalEntryCreator:
    def __init__(self):
        self.journal_lines = None
        self.grouped_entries = {}
        self.unassigned_lines = []
        self.additional_output_lines = []  # rows to append to output (e.g., plug lines)
        self._id_counts = {}
        self.workbook = None
        self.ws_jel = None
        self.ws_ctb = None
        self.ws_type_row = None
        self.ws_header_row = None
        self.ws_data_start_row = None
        self.template_type = 'old'
        self.skip_id_creation = False
        self._plug_lines_added = False

        # New template handling
        self.workbook = None
        self.ws_jel = None
        self.ws_ctb = None
        self.template_type = 'old'
        self.skip_id_creation = False
        self._plug_lines_added = False

    def _to_decimal(self, value):
        try:
            return Decimal(str(value))
        except Exception:
            return Decimal('0')

    def _sum_decimal(self, series):
        total = Decimal('0')
        for v in series.fillna(0).tolist():
            total += self._to_decimal(v)
        return total.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    def _sanitize_field_name(self, name):
        s = str(name).strip()
        return ''.join(ch if ch.isalnum() or ch in (' ', '_', '-') else '_' for ch in s).replace(' ', '_')

    def _normalize_header_token(self, token):
        if token is None:
            return ''
        text = str(token)
        text = text.replace('\u00A0', ' ')
        text = ' '.join(text.strip().split())
        return text

    def _canonical(self, token):
        return self._normalize_header_token(token).lower()

    def _find_header_row(self, df_all, required_names, max_scan_rows=10):
        """Find a row in df_all (header=None) that contains all required column names (case/space-insensitive)."""
        required = {self._canonical(n) for n in required_names}
        rows_to_scan = min(max_scan_rows, len(df_all))
        for r in range(rows_to_scan):
            row_vals = [self._normalize_header_token(v) for v in df_all.iloc[r].tolist()]
            canon = {self._canonical(v) for v in row_vals}
            if required.issubset(canon):
                # Build unique headers preserving original tokens (normalized for cleanliness)
                seen = {}
                headers = []
                for v in row_vals:
                    name = v if v != '' else 'Unnamed'
                    count = seen.get(name, 0)
                    unique = f"{name}_{count}" if count > 0 else name
                    seen[name] = count + 1
                    headers.append(unique)
                return r, headers
        return None, None

    def _format_value_token(self, value):
        if pd.isna(value):
            return 'NA'
        if isinstance(value, (pd.Timestamp, np.datetime64)):
            try:
                dt = pd.to_datetime(value)
                return dt.strftime('%Y%m%d')
            except Exception:
                return self._sanitize_field_name(str(value))
        if isinstance(value, (float, int, np.floating, np.integer)):
            return str(value).replace('.', '_')
        # strings / others
        return self._sanitize_field_name(str(value))

    def generate_journal_id(self, grouping_fields, group_key, max_len=100):
        # Normalize grouping_fields and group_key to aligned values
        fields = list(grouping_fields) if isinstance(grouping_fields, (list, tuple)) else [grouping_fields]
        if isinstance(group_key, tuple):
            values = list(group_key)
        else:
            values = [group_key]
        min_len = min(len(fields), len(values))
        tokens = [self._format_value_token(values[i]) for i in range(min_len)]
        if not tokens:
            base = 'ID'
        else:
            base = '-'.join(tokens)
        # Truncate to max_len
        if len(base) > max_len:
            base = base[:max_len]
        # Ensure uniqueness across entire run while respecting max_len
        count = self._id_counts.get(base, 0)
        self._id_counts[base] = count + 1
        if count == 0:
            return base
        suffix = f"__{count+1}"
        # If appending suffix exceeds max_len, trim base accordingly
        if len(base) + len(suffix) > max_len:
            base = base[: max_len - len(suffix)]
        return f"{base}{suffix}"
    
    def _normalize_and_deduplicate_columns(self, columns):
        """Return stripped, non-empty, unique column names by suffixing duplicates."""
        normalized = []
        for col in columns:
            name = '' if pd.isna(col) else str(col).strip()
            if name.lower() in ('', 'nan', 'none'):
                name = 'Unnamed'
            normalized.append(name)
        seen = {}
        unique_cols = []
        for name in normalized:
            count = seen.get(name, 0)
            unique_name = f"{name}_{count}" if count > 0 else name
            seen[name] = count + 1
            unique_cols.append(unique_name)
        return unique_cols
        
    def load_data(self, file_path):
        """Load journal lines from Excel file"""
        try:
            # Try to load as new template first (Journal Entries & Lines sheet)
            try:
                wb = load_workbook(file_path)
                if 'Journal Entries & Lines' in wb.sheetnames:
                    self.workbook = wb
                    self.ws_jel = wb['Journal Entries & Lines']
                    self.ws_ctb = wb['Comparative Trial Balances'] if 'Comparative Trial Balances' in wb.sheetnames else None
                    self.template_type = 'new'
                    # Read raw sheet to detect header row robustly
                    df_all = pd.read_excel(file_path, sheet_name='Journal Entries & Lines', header=None)
                    if len(df_all) == 0:
                        print("No data found in Journal Entries & Lines sheet.")
                        return False
                    req = ['Posted Date', 'Account ID', 'Debit Amount', 'Credit Amount']
                    header_row, headers = self._find_header_row(df_all, req, max_scan_rows=10)
                    if header_row is None:
                        print("Error: Required columns not found in Journal Entries & Lines sheet")
                        return False
                    # header_row is 0-based index where actual field names are found
                    # In new template: row 0 = Required/Optional, row 1 = field names, row 2+ = data
                    # But _find_header_row returns the row with field names (row 1, zero-indexed)
                    self.ws_type_row = header_row  # 0-based row with Required/Optional (row 1 in Excel)
                    self.ws_header_row = header_row + 1  # Field names row in Excel (1-based)
                    self.ws_data_start_row = header_row + 2  # Data starts here in Excel (1-based)
                    df = df_all.iloc[header_row+1:].copy()
                    df.columns = headers[:len(df.columns)]
                    # Normalize critical columns by canonical name lookup
                    # Build mapping from canonical -> actual name
                    canon_map = {self._canonical(col): col for col in df.columns}
                    try:
                        posted_col = canon_map[self._canonical('Posted Date')]
                        acct_col = canon_map[self._canonical('Account ID')]
                        debit_col = canon_map[self._canonical('Debit Amount')]
                        credit_col = canon_map[self._canonical('Credit Amount')]
                    except KeyError:
                        print("Error: Required columns not found after header normalization")
                        return False
                    # Rename to standard names for internal processing
                    df = df.rename(columns={posted_col: 'Posted Date', acct_col: 'Account ID', debit_col: 'Debit Amount', credit_col: 'Credit Amount'})
                    df = df.dropna(how='all')
                    df['Posted Date'] = pd.to_datetime(df['Posted Date'], errors='coerce').dt.normalize()
                    df['Debit Amount'] = pd.to_numeric(df['Debit Amount'], errors='coerce').fillna(0)
                    df['Credit Amount'] = pd.to_numeric(df['Credit Amount'], errors='coerce').fillna(0)
                    df = df.reset_index(drop=True)
                    df['_row_index'] = range(len(df))
                    self.journal_lines = df
                    print(f"Loaded {len(df)} journal lines from 'Journal Entries & Lines' in {file_path}")
                    return True
            except Exception:
                # Fall back to old template if anything fails
                pass

            # Read the entire file to analyze structure (old template)
            df_all = pd.read_excel(file_path, header=None)
            
            if len(df_all) == 0:
                print("No data found in Excel file.")
                return False
            
            # Check the structure - look for the actual field names
            first_row = df_all.iloc[0].values
            second_row = df_all.iloc[1].values if len(df_all) > 1 else None
            
            # Check different possible structures
            if second_row is not None and isinstance(second_row[0], str) and 'Posted Date' in str(second_row[0]):
                print("Detected field names in second row (row 1)")
                template_row = second_row
                df = df_all.iloc[2:].copy()  # Data starts from row 3 (index 2)
            elif isinstance(first_row[0], str) and 'Posted Date' in str(first_row[0]):
                print("Detected template headers in first row")
                template_row = first_row
                df = df_all.iloc[1:].copy()  # Data starts from row 2
            else:
                print("No template headers detected, assuming data starts from first row")
                # Use default column names
                template_row = ['Posted Date', 'Account ID', 'Debit Amount', 'Credit Amount'] + \
                              [f'Optional_{i}' for i in range(len(df_all.columns) - 4)]
                df = df_all.copy()
            
            if len(df) == 0:
                print("No data found in Excel file. Please add journal line data.")
                return False
            
            # Set column names from template and ensure uniqueness/cleanliness
            df.columns = template_row[:len(df.columns)]  # Handle case where data has fewer columns
            df.columns = self._normalize_and_deduplicate_columns(df.columns)
            
            # Remove completely empty rows
            df = df.dropna(how='all')
            
            # Ensure required columns exist and have data
            required_cols = ['Posted Date', 'Account ID', 'Debit Amount', 'Credit Amount']
            for col in required_cols:
                if col not in df.columns:
                    print(f"Error: Required column '{col}' not found in Excel file")
                    return False
            
            # Convert date column to datetime and normalize to date-only (00:00)
            df['Posted Date'] = pd.to_datetime(df['Posted Date'], errors='coerce')
            df['Posted Date'] = df['Posted Date'].dt.normalize()
            
            # Convert amount columns to numeric, filling NaN with 0
            df['Debit Amount'] = pd.to_numeric(df['Debit Amount'], errors='coerce').fillna(0)
            df['Credit Amount'] = pd.to_numeric(df['Credit Amount'], errors='coerce').fillna(0)
            
            # Validate that each line has either debit OR credit, not both
            both_amounts = (df['Debit Amount'] != 0) & (df['Credit Amount'] != 0)
            if both_amounts.any():
                print("Warning: Some lines have both debit and credit amounts. This may not be standard.")
            
            # Add row index for tracking
            df['_row_index'] = range(len(df))
            
            self.journal_lines = df
            print(f"Loaded {len(df)} journal lines from {file_path}")
            return True
            
        except Exception as e:
            print(f"Error loading Excel file: {e}")
            return False
    
    def get_optional_fields(self):
        """Get list of optional field column names that have data"""
        if self.journal_lines is None:
            return []
        
        # Find columns that are not required fields and have some non-null data
        required_cols = ['Posted Date', 'Account ID', 'Debit Amount', 'Credit Amount', '_row_index']
        optional_cols = []
        
        for col in self.journal_lines.columns:
            if col not in required_cols:
                # Check if column has any non-null, non-empty data
                try:
                    non_empty = self.journal_lines[col].notna() & \
                               (self.journal_lines[col] != '') & \
                               (self.journal_lines[col] != '[INSERT FIELD NAME]')
                    if non_empty.any():
                        optional_cols.append(col)
                except:
                    # Skip columns that cause issues
                    continue
        
        return optional_cols
    
    def validate_balances(self, epsilon=0.01, return_details=False):
        """Validate balances overall, by posted date, and by month.

        If return_details is True, returns a dict with diagnostics:
            {
              'ok': bool,
              'overall': {'debits': float, 'credits': float, 'net': float},
              'by_date': DataFrame with debits, credits, net, diff,
              'unbalanced_dates': DataFrame subset,
              'by_month': DataFrame with debits, credits, net, diff,
              'unbalanced_months': DataFrame subset,
              'messages': [str, ...]
            }
        Otherwise, raises ValueError on failure, returns True on success.
        """
        if self.journal_lines is None or len(self.journal_lines) == 0:
            if return_details:
                return {'ok': False, 'messages': ["No journal lines loaded to validate."]}
            raise ValueError("No journal lines loaded to validate.")

        if 'Posted Date' not in self.journal_lines.columns:
            if return_details:
                return {'ok': False, 'messages': ["Required column 'Posted Date' is missing."]}
            raise ValueError("Required column 'Posted Date' is missing.")

        df = self.journal_lines.copy()
        details = {'messages': []}

        # Overall
        total_debits = self._sum_decimal(df['Debit Amount'])
        total_credits = self._sum_decimal(df['Credit Amount'])
        overall_net = (total_debits - total_credits)
        details['overall'] = {
            'debits': total_debits,
            'credits': total_credits,
            'net': overall_net,
        }
        if abs(overall_net) >= epsilon:
            details['messages'].append(
                f"Overall debits and credits do not balance. Total Debits: ${total_debits:,.2f}, Total Credits: ${total_credits:,.2f}, Difference: ${overall_net:,.2f}."
            )

        # Per date
        df['__date__'] = df['Posted Date'].dt.date
        by_date = df.groupby('__date__', dropna=False)[['Debit Amount', 'Credit Amount']].sum()
        by_date = by_date.rename(columns={'Debit Amount': 'debits', 'Credit Amount': 'credits'})
        by_date['debits'] = by_date['debits'].apply(lambda x: Decimal(str(x))).apply(lambda x: x.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
        by_date['credits'] = by_date['credits'].apply(lambda x: Decimal(str(x))).apply(lambda x: x.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
        by_date['net'] = by_date['debits'] - by_date['credits']
        by_date['diff'] = by_date['net'].abs()
        unbalanced_dates = by_date[by_date['diff'] >= epsilon]
        details['by_date'] = by_date
        details['unbalanced_dates'] = unbalanced_dates
        if len(unbalanced_dates) > 0:
            msg_lines = ["Unbalanced posted dates detected (debits != credits):"]
            for d, row in unbalanced_dates.iterrows():
                msg_lines.append(
                    f"  - {d}: Debits=${row['debits']:,.2f}, Credits=${row['credits']:,.2f}, Difference=${row['net']:,.2f}"
                )
            details['messages'].append("\n".join(msg_lines))

        # Per month (only meaningful to report if any date unbalanced or overall unbalanced)
        df['__month__'] = df['Posted Date'].dt.to_period('M')
        by_month = df.groupby('__month__', dropna=False)[['Debit Amount', 'Credit Amount']].sum()
        by_month = by_month.rename(columns={'Debit Amount': 'debits', 'Credit Amount': 'credits'})
        by_month['debits'] = by_month['debits'].apply(lambda x: Decimal(str(x))).apply(lambda x: x.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
        by_month['credits'] = by_month['credits'].apply(lambda x: Decimal(str(x))).apply(lambda x: x.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
        by_month['net'] = by_month['debits'] - by_month['credits']
        by_month['diff'] = by_month['net'].abs()
        unbalanced_months = by_month[by_month['diff'] >= epsilon]
        details['by_month'] = by_month
        details['unbalanced_months'] = unbalanced_months
        if len(unbalanced_months) > 0:
            msg_lines = ["Unbalanced posted months detected (debits != credits):"]
            for m, row in unbalanced_months.iterrows():
                msg_lines.append(
                    f"  - {m}: Debits=${row['debits']:,.2f}, Credits=${row['credits']:,.2f}, Difference=${row['net']:,.2f}"
                )
            details['messages'].append("\n".join(msg_lines))

        # Cleanup
        df.drop(columns=['__date__', '__month__'], inplace=True)

        ok = len(details['messages']) == 0
        details['ok'] = ok
        if return_details:
            return details
        if not ok:
            raise ValueError("\n\n".join(details['messages']))
        return True

    def add_plug_lines_for_imbalances(self, details, plug_account_id="Audit Sight Clearing", epsilon=0.01):
        """Append plug lines to self.journal_lines to fix overall/date/month imbalances.

        Strategy:
        - If per-date imbalances exist: add one plug per unbalanced date on that date
        - Else if per-month imbalances exist: add one plug per unbalanced month on the last date present in that month
        - Else if only overall imbalance exists: add one plug dated on the latest Posted Date in data
        """
        if self.journal_lines is None or len(self.journal_lines) == 0:
            return 0

        df = self.journal_lines
        next_row_index = int(df['_row_index'].max()) + 1 if '_row_index' in df.columns and len(df) > 0 else 0
        plugs_added = 0

        def append_plug_row(date_value, net):
            nonlocal next_row_index, plugs_added
            if abs(net) < epsilon:
                return
            plug_debit = 0.0
            plug_credit = 0.0
            if net > 0:
                # More debits → add credit
                plug_credit = abs(net)
            else:
                plug_debit = abs(net)
            plug = {
                'Posted Date': pd.to_datetime(date_value),
                'Account ID': plug_account_id,
                'Debit Amount': plug_debit,
                'Credit Amount': plug_credit,
                '_row_index': next_row_index,
            }
            # Fill any other optional columns with None
            for col in df.columns:
                if col not in plug:
                    plug[col] = None
            self.journal_lines = pd.concat([self.journal_lines, pd.DataFrame([plug])], ignore_index=True)
            next_row_index += 1
            plugs_added += 1
            self._plug_lines_added = True
            # Also track for output appending
            add_line = {col: None for col in self.journal_lines.columns if col != '_row_index'}
            add_line['Posted Date'] = pd.to_datetime(date_value)
            add_line['Account ID'] = plug_account_id
            add_line['Debit Amount'] = plug_debit
            add_line['Credit Amount'] = plug_credit
            self.additional_output_lines.append(add_line)

        # Prefer date-level fix when present
        if details.get('unbalanced_dates') is not None and len(details['unbalanced_dates']) > 0:
            for date_value, row in details['unbalanced_dates'].iterrows():
                append_plug_row(date_value, float(row['net']))
            return plugs_added

        # Otherwise fix at month level
        if details.get('unbalanced_months') is not None and len(details['unbalanced_months']) > 0:
            # pick the last date present in each month
            df_tmp = self.journal_lines.copy()
            df_tmp['__month__'] = df_tmp['Posted Date'].dt.to_period('M')
            for month_period, row in details['unbalanced_months'].iterrows():
                # latest date in that month
                candidates = df_tmp[df_tmp['__month__'] == month_period]['Posted Date']
                if len(candidates) > 0:
                    date_value = pd.to_datetime(candidates.max())
                else:
                    # fallback: first day of month
                    date_value = pd.Period(month_period, freq='M').to_timestamp(how='end')
                append_plug_row(date_value, float(row['net']))
            return plugs_added

        # Otherwise only overall imbalance
        net = details.get('overall', {}).get('net', 0.0)
        if abs(net) >= epsilon:
            # use latest date in data
            last_date = pd.to_datetime(self.journal_lines['Posted Date'].max())
            append_plug_row(last_date, float(net))
        return plugs_added

    def check_balance(self, group_df):
        """Check if a group of journal lines is balanced and has minimum 2 lines"""
        # Must have at least 2 lines for a valid journal entry
        if len(group_df) < 2:
            return False
            
        total_debits = group_df['Debit Amount'].sum()
        total_credits = group_df['Credit Amount'].sum()
        
        # Use small epsilon for floating point comparison
        return abs(total_debits - total_credits) < 0.01
    
    def generate_grouping_combinations(self, optional_fields, max_fields=3):
        """Generate combinations of fields to try for grouping - prioritize more specific groupings first"""
        combinations_to_try = []
        
        # Start with most specific combinations (more fields = smaller groups)
        # Work backwards from max fields to 1 field, then date-only last
        for r in range(min(len(optional_fields), max_fields), 0, -1):
            for combo in combinations(optional_fields, r):
                combinations_to_try.append(['Posted Date'] + list(combo))
        
        # Add date-only grouping as the last resort
        combinations_to_try.append(['Posted Date'])
        
        return combinations_to_try
    
    def create_journal_entries(self, max_optional_fields=5):
        """Create journal entries by grouping lines that balance"""
        if self.journal_lines is None:
            print("No data loaded. Please load data first.")
            return False
        
        print("Creating journal entries...")
        
        # Get optional fields that have data
        optional_fields = self.get_optional_fields()
        print(f"Found optional fields with data: {optional_fields}")
        
        # Generate combinations of fields to try for grouping
        field_combinations = self.generate_grouping_combinations(optional_fields, max_optional_fields)
        print(f"Will try {len(field_combinations)} different grouping combinations")
        
        # Track which lines have been assigned to journal entries
        assigned_lines = set()
        journal_entry_id = 1
        
        # Try each combination of grouping fields
        for fields in field_combinations:
            print(f"\nTrying grouping by: {fields}")
            
            # Get unassigned lines using a safer approach
            unassigned_mask = ~self.journal_lines['_row_index'].isin(assigned_lines)
            unassigned_df = self.journal_lines[unassigned_mask].copy()
            
            if len(unassigned_df) == 0:
                print("All lines have been assigned to journal entries.")
                break
            
            # Reset index to avoid any potential conflicts
            unassigned_df = unassigned_df.reset_index(drop=True)
            
            # Group by the current field combination
            valid_fields = [col for col in fields if col in unassigned_df.columns]
            # Dedupe fields and guard against empty grouping list
            valid_fields = list(dict.fromkeys(valid_fields))
            if not valid_fields:
                continue
            
            # Use a more robust grouping approach
            try:
                # Create a safe copy for grouping
                grouping_df = unassigned_df[valid_fields + ['Debit Amount', 'Credit Amount', '_row_index']].copy()
                grouped = grouping_df.groupby(valid_fields, dropna=False, sort=False)
            except Exception as e:
                print(f"   Error grouping by {valid_fields}: {e}")
                continue
            
            balanced_groups = 0
            groups_processed = 0
            
            for group_key, group_indices in grouped.groups.items():
                groups_processed += 1
                # Get the actual group data using safe indexing
                group_df = unassigned_df.iloc[group_indices].copy()
                
                if self.check_balance(group_df):
                    # Create field-based journal entry ID
                    je_id = self.generate_journal_id(valid_fields, group_key)
                    
                    # Store the journal entry with completely fresh index
                    group_df_clean = group_df.copy()
                    group_df_clean = group_df_clean.reset_index(drop=True)
                    
                    self.grouped_entries[je_id] = {
                        'lines': group_df_clean,
                        'grouping_fields': fields,
                        'group_key': group_key,
                        'total_debits': group_df_clean['Debit Amount'].sum(),
                        'total_credits': group_df_clean['Credit Amount'].sum()
                    }
                    
                    # Mark these lines as assigned using original row indices
                    original_indices = group_df['_row_index'].tolist()
                    assigned_lines.update(original_indices)
                    
                    balanced_groups += 1
            
            print(f"Found {balanced_groups} balanced journal entries from {groups_processed} groups with this grouping")
            
            # If this grouping level found balanced entries, continue with this level
            # to find more specific entries before moving to broader groupings
            if balanced_groups > 0:
                print(f"  -> Successfully created {balanced_groups} specific journal entries")
        
        # Handle remaining unassigned lines using safer approach
        remaining_mask = ~self.journal_lines['_row_index'].isin(assigned_lines)
        remaining_lines = self.journal_lines[remaining_mask].copy().reset_index(drop=True)
        
        if len(remaining_lines) > 0:
            print(f"\nProcessing {len(remaining_lines)} remaining lines...")
            
            for idx, line in remaining_lines.iterrows():
                debit = line['Debit Amount']
                credit = line['Credit Amount']
                line_date = line['Posted Date']
                original_row_idx = line['_row_index']
                
                # Handle zero-amount lines by assigning to existing journal entry on same date
                if debit == 0 and credit == 0:
                    # Try to attach zero-amount line to an existing JE on same date; otherwise leave unassigned
                    assigned_to_existing = False
                    for je_id, entry_data in self.grouped_entries.items():
                        if len(entry_data['lines']) > 0:
                            entry_date = entry_data['lines']['Posted Date'].iloc[0]
                            if entry_date.date() == line_date.date():
                                line_df = pd.DataFrame([line.to_dict()]).reset_index(drop=True)
                                existing_lines = entry_data['lines'].copy().reset_index(drop=True)
                                entry_data['lines'] = pd.concat([existing_lines, line_df], ignore_index=True)
                                entry_data['total_debits'] += line['Debit Amount']
                                entry_data['total_credits'] += line['Credit Amount']
                                assigned_lines.add(original_row_idx)
                                assigned_to_existing = True
                                print(f"   → Assigned zero-amount line to {je_id}")
                                break
                    if not assigned_to_existing:
                        # Keep unassigned; will be handled by plug balancing step
                        continue
                
                # Skip invalid lines (both debit and credit non-zero)
                elif debit != 0 and credit != 0:
                    print(f"   Invalid line (both debit and credit): Account {line['Account ID']}")
                    continue
                
                # Create individual journal entry for valid single lines
                else:
                    # Do not create single-line entries; leave for plug balancing
                    continue
        
        # Track any remaining lines as unassigned (to be balanced or reported)
        self.unassigned_lines = self.journal_lines[~self.journal_lines['_row_index'].isin(assigned_lines)].copy()
        
        print(f"\nSummary:")
        print(f"Total journal entries created: {len(self.grouped_entries)}")
        print(f"Total lines assigned: {len(assigned_lines)}")
        print(f"Total lines unassigned (invalid): {len(self.unassigned_lines)}")
        
        return True
    
    def balance_unassigned_with_plug(self, plug_account_id="Audit Sight Clearing", epsilon=0.01):
        """Create balancing journal entries per posted date for unassigned lines by adding a plug line.

        - Groups unassigned lines by Posted Date (date-only)
        - For each date group, creates a new JE containing those lines
        - Adds one plug line with Account ID = plug_account_id for the offset amount
        - Records the plug line to be appended to output
        Returns number of dates balanced (int).
        """
        if self.journal_lines is None:
            print("No data loaded. Cannot balance.")
            return 0
        if len(self.unassigned_lines) == 0:
            print("No unassigned lines to balance.")
            return 0

        dates_balanced = 0
        # Determine next journal entry id number
        next_id_num = 1
        if len(self.grouped_entries) > 0:
            # existing IDs like JE0001
            try:
                nums = [int(k[2:]) for k in self.grouped_entries.keys() if str(k).startswith("JE")]
                if nums:
                    next_id_num = max(nums) + 1
            except Exception:
                pass

        # Group unassigned by date
        df_un = self.unassigned_lines.copy()
        df_un['__date__'] = df_un['Posted Date'].dt.date
        for date_value, group in df_un.groupby('__date__'):
            # Build new JE lines from original unassigned lines
            group_clean = group.copy().reset_index(drop=True)
            total_debits = self._sum_decimal(group_clean['Debit Amount'])
            total_credits = self._sum_decimal(group_clean['Credit Amount'])
            net = total_debits - total_credits

            je_id = self.generate_journal_id(['Posted Date'], (date_value,))

            # Create plug line if needed
            plug_debit = Decimal('0.00')
            plug_credit = Decimal('0.00')
            if net != Decimal('0.00'):
                if net > 0:
                    plug_credit = net.copy_abs().quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                else:
                    plug_debit = net.copy_abs().quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

                # Build plug line with same schema as journal_lines
                plug_row = {col: None for col in self.journal_lines.columns}
                plug_row['Posted Date'] = pd.to_datetime(date_value)
                plug_row['Account ID'] = plug_account_id
                plug_row['Debit Amount'] = float(plug_debit)
                plug_row['Credit Amount'] = float(plug_credit)
                plug_row['_row_index'] = -1  # not in original data

                # Append plug row into the JE lines for internal reporting
                group_with_plug = pd.concat([group_clean, pd.DataFrame([plug_row])], ignore_index=True)
            else:
                group_with_plug = group_clean

            # Store as a new JE
            self.grouped_entries[je_id] = {
                'lines': group_with_plug.reset_index(drop=True),
                'grouping_fields': ['Posted Date'],
                'group_key': str(date_value),
                'total_debits': float(group_with_plug['Debit Amount'].sum()),
                'total_credits': float(group_with_plug['Credit Amount'].sum())
            }

            # Mark original unassigned as now assigned for output mapping
            # (we'll assign Journal IDs via _row_index mapping; plug line handled separately)

            # Record additional output line for plug (without _row_index)
            if net != Decimal('0.00'):
                add_line = {col: None for col in self.journal_lines.columns if col != '_row_index'}
                add_line['Posted Date'] = pd.to_datetime(date_value)
                add_line['Account ID'] = plug_account_id
                add_line['Debit Amount'] = float(plug_debit)
                add_line['Credit Amount'] = float(plug_credit)
                # Keep for output append with Journal ID later
                self.additional_output_lines.append({'Journal ID': je_id, **add_line})
                self._plug_lines_added = True

            dates_balanced += 1

        # After creating new JEs, clear unassigned_lines so reporting reflects resolution
        self.unassigned_lines = self.journal_lines.iloc[0:0].copy()
        return dates_balanced
    
    def generate_output(self, input_file_path, output_file_path=None):
        """Generate output Excel file with Journal ID column"""
        if self.journal_lines is None:
            print("No data to output")
            return False
        
        # If using new template workbook, preserve sheets and write IDs/plug lines there
        if self.template_type == 'new' and self.workbook is not None and self.ws_jel is not None:
            # Determine output path
            if output_file_path is None:
                input_path = Path(input_file_path)
                output_file_path = input_path.parent / f"{input_path.stem}_with_journal_ids{input_path.suffix}"

            # If skip flag set, just save
            if self.skip_id_creation:
                try:
                    self.workbook.save(output_file_path)
                    print(f"Output written to: {output_file_path}")
                    self.print_summary_report()
                    return True
                except Exception as e:
                    print(f"Error writing output file: {e}")
                    return False

            # Otherwise assign Journal IDs - ALWAYS write to column A (column 1)
            jid_col_idx = 1
            # Set header in column A
            self.ws_jel.cell(row=self.ws_header_row, column=jid_col_idx, value='Journal ID')
            # Build complete header list for plug line mapping
            header_cells = self.ws_jel[self.ws_header_row]
            headers = [cell.value for cell in header_cells]
            
            # Find the last row with actual data (not just Journal IDs we might write)
            last_data_row = self.ws_data_start_row - 1
            for row in self.ws_jel.iter_rows(min_row=self.ws_data_start_row):
                # Check if any cell in this row (excluding column A) has data
                has_data = any(cell.value is not None for cell in row[1:])  # Skip column A
                if has_data:
                    last_data_row = row[0].row
                else:
                    break
            
            # Write Journal IDs to column A for all assigned lines (skip plug lines with _row_index < 0)
            for je_id, entry_data in self.grouped_entries.items():
                row_indices = entry_data['lines']['_row_index'].tolist()
                for idx in row_indices:
                    # Skip plug lines that were added during balancing (they have negative row index)
                    if int(idx) < 0:
                        continue
                    ws_row = self.ws_data_start_row + int(idx)
                    self.ws_jel.cell(row=ws_row, column=jid_col_idx, value=je_id)
            
            # Append plug lines to the bottom of the sheet - start right after last data row
            if self.additional_output_lines:
                # Map column names to indices
                col_to_idx = {}
                for i, h in enumerate(headers):
                    if h:
                        col_to_idx[h] = i + 1
                # Also need to map our internal column names to sheet columns
                # Find Posted Date, Account ID, Debit Amount, Credit Amount columns
                for i, h in enumerate(headers):
                    if h and 'posted' in str(h).lower() and 'date' in str(h).lower():
                        col_to_idx['Posted Date'] = i + 1
                    elif h and 'account' in str(h).lower() and 'id' in str(h).lower():
                        col_to_idx['Account ID'] = i + 1
                    elif h and 'debit' in str(h).lower():
                        col_to_idx['Debit Amount'] = i + 1
                    elif h and 'credit' in str(h).lower():
                        col_to_idx['Credit Amount'] = i + 1
                
                # Write each plug line row starting from last_data_row + 1
                plug_row_start = last_data_row + 1
                for idx, plug_line in enumerate(self.additional_output_lines):
                    new_row_idx = plug_row_start + idx
                    # Write Journal ID if present in plug_line dict
                    if 'Journal ID' in plug_line:
                        self.ws_jel.cell(row=new_row_idx, column=jid_col_idx, value=plug_line['Journal ID'])
                    # Write other fields
                    for key, val in plug_line.items():
                        if key == 'Journal ID':
                            continue
                        if key in col_to_idx:
                            self.ws_jel.cell(row=new_row_idx, column=col_to_idx[key], value=val)
            # Add clearing account to CTB if plug lines were added
            if self._plug_lines_added and self.ws_ctb is not None:
                ctbr = self.ws_ctb.max_row + 1
                self.ws_ctb.cell(row=ctbr, column=1, value='Audit Sight Clearing')
                self.ws_ctb.cell(row=ctbr, column=2, value='Audit Sight Clearing')
                self.ws_ctb.cell(row=ctbr, column=3, value=0)
                self.ws_ctb.cell(row=ctbr, column=4, value=0)
                self.ws_ctb.cell(row=ctbr, column=5, value='Assets')
                self.ws_ctb.cell(row=ctbr, column=6, value='asset:current:other')
                print("Audit Sight Clearing account added to the trial balance in the Comparative Trial Balances tab.")
            try:
                self.workbook.save(output_file_path)
                print(f"Output written to: {output_file_path}")
                self.print_summary_report()
                return True
            except Exception as e:
                print(f"Error writing output file: {e}")
                return False

        # Fallback: original single-sheet export
        output_df = self.journal_lines.copy()
        output_df['Journal ID'] = ''
        for je_id, entry_data in self.grouped_entries.items():
            row_indices = entry_data['lines']['_row_index'].tolist()
            output_df.loc[output_df['_row_index'].isin(row_indices), 'Journal ID'] = je_id
        output_df = output_df.drop('_row_index', axis=1)
        if self.additional_output_lines:
            add_rows = []
            for row in self.additional_output_lines:
                completed = {col: row.get(col, None) for col in output_df.columns}
                add_rows.append(completed)
            if add_rows:
                output_df = pd.concat([output_df, pd.DataFrame(add_rows)], ignore_index=True)
        cols = list(output_df.columns)
        cols.remove('Journal ID')
        posted_date_idx = cols.index('Posted Date')
        cols.insert(posted_date_idx, 'Journal ID')
        output_df = output_df[cols]
        if output_file_path is None:
            input_path = Path(input_file_path)
            output_file_path = input_path.parent / f"{input_path.stem}_with_journal_ids{input_path.suffix}"
        try:
            output_df.to_excel(output_file_path, index=False)
            print(f"Output written to: {output_file_path}")
            self.print_summary_report()
            return True
        except Exception as e:
            print(f"Error writing output file: {e}")
            return False
    
    def print_summary_report(self):
        """Print a summary report of journal entries created"""
        print("\n" + "="*60)
        print("JOURNAL ENTRY SUMMARY REPORT")
        print("="*60)
        
        # Separate multi-line and single-line entries for better reporting
        multi_line_entries = []
        single_line_entries = []
        
        for je_id, entry_data in sorted(self.grouped_entries.items()):
            if len(entry_data['lines']) > 1:
                multi_line_entries.append((je_id, entry_data))
            else:
                single_line_entries.append((je_id, entry_data))
        
        # Report multi-line entries
        if multi_line_entries:
            print(f"\nMULTI-LINE JOURNAL ENTRIES ({len(multi_line_entries)}):")
            for je_id, entry_data in multi_line_entries:
                lines = entry_data['lines']
                print(f"\n{je_id}:")
                print(f"  Date: {lines['Posted Date'].iloc[0].strftime('%Y-%m-%d')}")
                print(f"  Lines: {len(lines)}")
                print(f"  Total Debits: ${entry_data['total_debits']:,.2f}")
                print(f"  Total Credits: ${entry_data['total_credits']:,.2f}")
                print(f"  Grouped by: {entry_data['grouping_fields']}")
                if len(entry_data['grouping_fields']) > 1 and entry_data['grouping_fields'][0] != 'Individual Entry':
                    print(f"  Group values: {entry_data['group_key']}")
        
        # Report single-line entries (summary only)
        if single_line_entries:
            print(f"\nSINGLE-LINE JOURNAL ENTRIES ({len(single_line_entries)}):")
            print("  (Each line created as individual journal entry)")
            for je_id, entry_data in single_line_entries[:5]:  # Show first 5 examples
                lines = entry_data['lines']
                line = lines.iloc[0]
                print(f"  {je_id}: {line['Posted Date'].strftime('%Y-%m-%d')}, "
                      f"Account: {line['Account ID']}, "
                      f"Debit: ${line['Debit Amount']:.2f}, "
                      f"Credit: ${line['Credit Amount']:.2f}")
            
            if len(single_line_entries) > 5:
                print(f"  ... and {len(single_line_entries) - 5} more single-line entries")
        
        if len(self.unassigned_lines) > 0:
            print(f"\nINVALID LINES ({len(self.unassigned_lines)}):")
            print("  (Lines with both debit and credit, or both zero)")
            for _, line in self.unassigned_lines.iterrows():
                print(f"  Date: {line['Posted Date'].strftime('%Y-%m-%d')}, "
                      f"Account: {line['Account ID']}, "
                      f"Debit: ${line['Debit Amount']:.2f}, "
                      f"Credit: ${line['Credit Amount']:.2f}")

def main():
    parser = argparse.ArgumentParser(description='Create balanced journal entries from Excel file')
    parser.add_argument('input_file', help='Path to input Excel file')
    parser.add_argument('-o', '--output', help='Path to output Excel file (optional)')
    parser.add_argument('--max-fields', type=int, default=5, 
                       help='Maximum number of optional fields to use for grouping (default: 5)')
    parser.add_argument('--auto-balance', action='store_true',
                       help="If set, automatically add plug lines using 'Audit Sight Clearing' to balance unassigned dates")
    
    args = parser.parse_args()
    
    # Create journal entry creator
    creator = JournalEntryCreator()
    
    try:
        # Load data
        if not creator.load_data(args.input_file):
            return 1
        
        # If new template with Journal ID column, consider skipping creation
        if creator.evaluate_provided_journal_ids():
            # Directly proceed to output with original workbook
            if not creator.generate_output(args.input_file, args.output):
                return 1
            return 0
        else:
            # Validate balances (overall, per-date, and per-month when needed)
            details = creator.validate_balances(return_details=True)
            if not details['ok']:
                print("Imbalances detected:\n" + "\n\n".join(details['messages']))
                print("Unbalanced journal entries or journal entries on multiple dates detected. Running journal entry ID creation logic.")
                if args.auto_balance:
                    added = creator.add_plug_lines_for_imbalances(details)
                    print(f"Auto-added {added} plug line(s) to balance.")
                else:
                    try:
                        resp = input("Add plug lines using 'Audit Sight Clearing' to fix these imbalances? [y/N]: ").strip().lower()
                    except EOFError:
                        resp = 'n'
                    if resp == 'y':
                        added = creator.add_plug_lines_for_imbalances(details)
                        print(f"Added {added} plug line(s) to balance.")
                    else:
                        print("User declined to auto-balance imbalances. Aborting.")
                        return 1
        
        # Create journal entries
        if not creator.create_journal_entries(args.max_fields):
            return 1
        
        # If unassigned lines remain, optionally prompt in CLI or auto-balance via flag
        if len(creator.unassigned_lines) > 0:
            print(f"Unassigned lines remain: {len(creator.unassigned_lines)}")
            if args.auto_balance:
                balanced_dates = creator.balance_unassigned_with_plug()
                print(f"Auto-balanced by adding plug lines for {balanced_dates} posted date(s).")
            else:
                # Interactive prompt
                try:
                    resp = input("Add plug lines using 'Audit Sight Clearing' to balance unassigned dates? [y/N]: ").strip().lower()
                except EOFError:
                    resp = 'n'
                if resp == 'y':
                    balanced_dates = creator.balance_unassigned_with_plug()
                    print(f"Added plug lines for {balanced_dates} posted date(s).")
                else:
                    print("User declined to auto-balance unassigned lines. Exiting.")
                    return 1
        
        # Generate output
        if not creator.generate_output(args.input_file, args.output):
            return 1
        
        return 0
    except Exception as e:
        print(f"Error: {e}")
        return 1

if __name__ == "__main__":
    exit(main())
